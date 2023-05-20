from twilio.rest import Client
from datetime import date
from openpyxl import load_workbook

class MobileAnnouncer():
    global phone_number
    phone_number = "+12707389977"

    def __init__(self, list_of_nums):
        
        self.list_of_nums = list_of_nums

        account = "AC3785333a26e0602c665cda119a72a594"
        token = "e32cb19a65c6cf3ddf469a2f02b85cda"
        self.client = Client(account, token)
        today = date.today()
        self.day_today = today.strftime("%d/%m/%Y")

    def normalMessage(self, body_message):
        self.body_message = body_message

        for num in self.list_of_nums:
            message = self.client.messages.create(
                body=f"{self.day_today}\n\n{self.body_message}\n\nThank You,\nOasis Community Hub Hobmoor",
                from_=phone_number,
                to={num}
            )

    def importantAnnouncement(self, body_message):
        self.body_message = body_message

        for num in self.list_of_nums:
            message = self.client.messages.create(
                body=f"{self.day_today}\n\nIMPORTANT ANNOUNCEMENT\n\n{self.body_message}\n\nThank You,\nOasis Community Hub Hobmoor",
                from_=phone_number,
                to={num}
            )
        

    def eventAnnouncement(self, body_message):
        self.body_message = body_message

        for num in self.list_of_nums:
            message = self.client.messages.create(
                body=f"{self.day_today}\n\nEVENT\n\n{self.body_message}\n\nThank You,\nOasis Community Hub Hobmoor",
                from_=phone_number,
                to={num}
            )

class excelScraper():

    def __init__(self, file_dir):

        self.file_dir = file_dir

        self.wb = load_workbook(file_dir)
        self.ws = self.wb.active

    def retrieveNumbers(self):
        column_data = self.ws["A"]
        raw_list_of_numbers = [column_data[i].value for i in range(len(column_data))]
        list_of_numbers = [("+44" + str(i)) for i in raw_list_of_numbers if isinstance(i, int)]
        return list_of_numbers

        


        
        
        



        


